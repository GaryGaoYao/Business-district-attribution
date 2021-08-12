import PositionDesition
import openpyxl
import numpy as np
import matplotlib.pyplot as plt

#商圈数据
workbook1 = openpyxl.load_workbook('D:\\IE-Download\\shangquanzuobiao-0701.xlsx')
worksheet1 = workbook1.active
minrow1 = worksheet1.min_row #最小行
maxrow1 = worksheet1.max_row #最大行

pointlist_market = []
namelist = []
partitionlist = []
partitionlist_jingweidu = []

#164个商圈
for n in range(minrow1, maxrow1+1):
    cell_count = worksheet1.cell(n, 2).value
    if cell_count == 0:
        namelist.append(worksheet1.cell(n, 1).value)
        partitionlist.append(n-1)
        #隔断的经纬度
        cell_jingdu = worksheet1.cell(n, 3).value
        cell_weidu = worksheet1.cell(n, 4).value

#经纬度度读取
for n in range(minrow1, maxrow1+1):
    cell_jingdu = worksheet1.cell(n, 3).value
    cell_weidu = worksheet1.cell(n, 4).value
    pointlist_market.append((cell_jingdu, cell_weidu))

print(pointlist_market)
print(namelist)
print(partitionlist)





