import PositionDesition
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
import shangquanduqu

workbook1 = openpyxl.load_workbook('G:\\python+codes\position+zhangxiyu\\xiyu\\tr-0812.xlsx')
worksheet1 = workbook1.active
minrow1 = worksheet1.min_row  # 最小行
maxrow1 = worksheet1.max_row  # 最大行

pointlist_market = []
PointListMap = []
shangquan_temporary_list = []
x = []
y = []

# 全量商户的经纬度度读取
for n in range(minrow1, maxrow1 + 1):
    cell_jingdu = worksheet1.cell(n, 18).value
    cell_weidu = worksheet1.cell(n, 19).value
    pointlist_market.append((cell_jingdu, cell_weidu))

for n in range(1, 164):
    shangquan_name_value = shangquanduqu.namelist[n-1]
    shangquan_temporary_list_start = shangquanduqu.partitionlist[n-1]
    shangquan_temporary_list_end = shangquanduqu.partitionlist[n]

    for i in range(shangquanduqu.partitionlist[n-1],shangquanduqu.partitionlist[n]):
        shangquan_temporary_list.append(shangquanduqu.pointlist_market[i])
        if i == shangquanduqu.partitionlist[n]-1:
            #判断全量商户坐标
            Insert_column_number = 60
            for j in range(0, len(pointlist_market)):
                a_x, a_y = pointlist_market[j]
                if (PositionDesition.IsPtInPoly(a_x, a_y, shangquan_temporary_list)):
                    worksheet1.cell(j + 1, Insert_column_number).value = shangquan_name_value

    #清空零时列表
    shangquan_temporary_list = []

#最后一个补充判断：贵友大厦通州店
guizhoudasha_tongzhou = [(116.672887,39.891627),(116.673812,39.891059),(116.673237,39.889253),(116.672087,39.889474)]
for j in range(0, len(pointlist_market)):
    a_x, a_y = pointlist_market[j]
    if (PositionDesition.IsPtInPoly(a_x, a_y, guizhoudasha_tongzhou)):
        worksheet1.cell(j + 1, 60).value = '贵友大厦通州店'

# 网格坐标区域（必须要顺时针顺序）
#示例
'''
PointListMap_shengwuyiyaojidi = [(116.322153, 39.697358), (116.335808, 39.697025), (116.334227, 39.69014),
                                 (116.33128, 39.675869), (116.322728, 39.675703)]
Insert_column_number = 50
for n in range(0, len(pointlist_market)):
    a_x, a_y = pointlist_market[n]

    if (PositionDesition.IsPtInPoly(a_x, a_y, PointListMap_shengwuyiyaojidi)):
        worksheet1.cell(n + 1, Insert_column_number).value = '生物医药基地'

    elif (PositionDesition.IsPtInPoly(a_x, a_y, PointListMap_fengtaikejiyuan)):
        worksheet1.cell(n + 1, Insert_column_number).value = '丰台科技园'

    elif (PositionDesition.IsPtInPoly(a_x, a_y, PointListMap_wukesong)):
        worksheet1.cell(n + 1, Insert_column_number).value = '五棵松'

    elif (PositionDesition.IsPtInPoly(a_x, a_y, PointListMap_dazhongshi)):
        worksheet1.cell(n + 1, Insert_column_number).value = '大钟寺'

    elif (PositionDesition.IsPtInPoly(a_x, a_y, PointListMap_yuandalu)):
        worksheet1.cell(n + 1, Insert_column_number).value = '远大路'

    elif (PositionDesition.IsPtInPoly(a_x, a_y, PointListMap_zhongguanchun)):
        worksheet1.cell(n + 1, Insert_column_number).value = '中关村'

    elif (PositionDesition.IsPtInPoly(a_x, a_y, PointListMap_wudaokou)):
        worksheet1.cell(n + 1, Insert_column_number).value = '五道口'

    elif (PositionDesition.IsPtInPoly(a_x, a_y, PointListMap_aoyuncun)):
        worksheet1.cell(n + 1, Insert_column_number).value = '奥运村'

    elif (PositionDesition.IsPtInPoly(a_x, a_y, PointListMap_changying)):
        worksheet1.cell(n + 1, Insert_column_number).value = '常营'

    else:
        worksheet1.cell(n + 1, Insert_column_number).value = '无'
'''

'''
for n in range(minrow2,maxrow2+1):
    cell_map_jingdu = sheets[1].cell(n, 1).value
    cell_map_weidu = sheets[1].cell(n, 2).value
    x.append(cell_map_jingdu)
    y.append(cell_map_weidu)

x.append(sheets[1].cell(1, 1).value)
y.append(sheets[1].cell(1, 2).value)

'''

workbook1.save('tr-0812_result.xlsx')
